# bot_utils.py
from typing import List
from models import Message, Conversation
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from datetime import datetime


LANG_MAP = {
    "ru": "Русский",
    "kk": "Казахский",
}


def export_conversations_to_excel(conversations: List[Conversation], period: str, file_path: str = "conversations.xlsx") -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Диалоги"

    headers = [
        "Язык",
        "Кто",
        "Сообщение",
        "Дата и время"
    ]
    
    bold_font = Font(bold=True)
    cell1 = ws.cell(row=1, column=1, value=f"Отзывы за период {period}")
    cell1.font = bold_font

    for conv_idx, conv in enumerate(conversations, start=1):
        last_m = conv.messages[-1]

        # --- Заголовок диалога ---
        title_row = ws.max_row + 3
        cell1 = ws.cell(row=title_row, column=1, value=f"Чат №{conv_idx}")
        cell1.font = bold_font
        cell2 = ws.cell(row=title_row, column=2, value=f"Клиент: {conv.client_phone}")
        cell2.font = bold_font

        # --- Пустая строка ---
        ws.append([])

        # --- Шапка таблицы ---
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = bold_font

        # --- Сообщения ---
        for msg in conv.messages:
            who = "Клиент" if msg.from_phone == conv.client_phone else "Бот"
            lang = LANG_MAP.get(last_m.language, last_m.language)  # преобразуем код языка

            ws.append([
                lang,
                who,
                msg.text,
                clean_datetime(msg.date_time),
            ])

            # Перенос текста в строке
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # --- Две пустые строки после каждого диалога ---
        ws.append([])
        ws.append([])

    # --- Автоширина колонок ---
    for i, col in enumerate(ws.columns, start=1):
        max_length = 0
        col_letter = get_column_letter(i)
        for cell in col:
            if cell.value:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        ws.column_dimensions[col_letter].width = min(max_length + 2, 120)

    # --- Автовысота строк для длинных сообщений ---
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        text_cell = row[2]  # колонка "Сообщение" (3-я)
        if text_cell.value:
            text = str(text_cell.value)
            line_count = text.count("\n") + 1
            wrap_factor = len(text) // 80 + 1
            ws.row_dimensions[text_cell.row].height = max(20, 15 * max(line_count, wrap_factor))

    wb.save(file_path)
    return file_path


def clean_datetime(dt_str: str) -> str:
    """
    Конвертирует строки вида:
    - '2025-08-31T13:39:02+05:00'
    - '2025-08-31 13:39:14 +0500 +05'
    В формат: '2025-08-31 13:39:02'
    """
    try:
        dt = datetime.fromisoformat(dt_str)
    except ValueError:
        cleaned = dt_str.split("+")[0].strip()
        dt = datetime.strptime(cleaned, "%Y-%m-%d %H:%M:%S")
    
    return dt.strftime("%Y-%m-%d %H:%M:%S")